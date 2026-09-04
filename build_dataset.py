"""
Builds the clean, dashboard-ready dataset from the raw TFC exports and finance report.
Run once locally to produce /data/*.csv, which the Streamlit app reads at runtime.
"""
import openpyxl
import pandas as pd

SRC = '/mnt/user-data/uploads/TFC_0_6.xlsx'
FIN_SRC = '/mnt/user-data/uploads/FinanceReport__3_.xlsx'
OUT = '/home/claude/tfc_dashboard/data'


def sheet_df(wb, name):
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h else h for h in rows[0]]
    return pd.DataFrame(rows[1:], columns=header)


wb = openpyxl.load_workbook(SRC, data_only=True)

component = sheet_df(wb, 'Component')
supplier = sheet_df(wb, 'Supplier')
product = sheet_df(wb, 'Product')
customer_product = sheet_df(wb, 'Customer - Product')
warehouse = sheet_df(wb, 'Warehouse, Salesarea')
bottling = sheet_df(wb, 'Bottling line')

# ---------------- Finance report: reshape wide -> long ----------------
wbf = openpyxl.load_workbook(FIN_SRC, data_only=True)
wsf = wbf['Output']
rows = list(wsf.iter_rows(values_only=True))
rounds = list(rows[0][1:])
fin_rows = []
for r in rows[1:]:
    label = r[0]
    if label is None:
        continue
    for rnd, val in zip(rounds, r[1:]):
        fin_rows.append({'line_item': label, 'round': rnd, 'value': val})
finance_long = pd.DataFrame(fin_rows)


def fin_pick(label):
    sub = finance_long[finance_long['line_item'] == label][['round', 'value']].copy()
    sub = sub.rename(columns={'value': label})
    return sub


# Core financial KPIs, one row per round
roi = fin_pick('ROI')
revenue = fin_pick('Realized revenue')
purchase_value = fin_pick('Gross margin - Cost of goods sold - Purchase value')
production_costs = fin_pick('Gross margin - Cost of goods sold - Production costs')
cogs = fin_pick('Gross margin - Cost of goods sold')
indirect_cost = fin_pick('Operating profit - Indirect cost')
operating_profit = fin_pick('Operating profit')
investment = fin_pick('Investment')

financial_kpis = roi.merge(revenue, on='round') \
    .merge(purchase_value, on='round') \
    .merge(production_costs, on='round') \
    .merge(cogs, on='round') \
    .merge(indirect_cost, on='round') \
    .merge(operating_profit, on='round') \
    .merge(investment, on='round')
financial_kpis = financial_kpis.rename(columns={
    'ROI': 'roi',
    'Realized revenue': 'realized_revenue',
    'Gross margin - Cost of goods sold - Purchase value': 'purchase_value',
    'Gross margin - Cost of goods sold - Production costs': 'production_costs',
    'Gross margin - Cost of goods sold': 'cogs',
    'Operating profit - Indirect cost': 'indirect_cost',
    'Operating profit': 'operating_profit',
    'Investment': 'investment',
})
financial_kpis = financial_kpis.sort_values('round').reset_index(drop=True)

# Purchase value by supplier, long format (for the Purchasing dashboard)
supplier_purchase_rows = []
for line in finance_long['line_item'].unique():
    if line.startswith('Gross margin - Cost of goods sold - Purchase value - '):
        sup_name = line.split(' - ')[-1]
        sub = finance_long[finance_long['line_item'] == line][['round', 'value']].copy()
        sub['supplier'] = sup_name
        supplier_purchase_rows.append(sub)
supplier_purchase = pd.concat(supplier_purchase_rows, ignore_index=True)
supplier_purchase = supplier_purchase[supplier_purchase['value'].fillna(0) != 0]

# Bonus/penalty by customer, long format (for the Sales dashboard)
customer_bonus_rows = []
for cust in ['Food & Groceries', 'LAND Market', "Dominick's"]:
    line = f'Realized revenue - Bonus or penalties - Contracted sales revenue - {cust}'
    sub = finance_long[finance_long['line_item'] == line][['round', 'value']].copy()
    sub['customer'] = cust
    customer_bonus_rows.append(sub)
customer_bonus = pd.concat(customer_bonus_rows, ignore_index=True)
customer_bonus = customer_bonus.rename(columns={'value': 'bonus_penalty'})

customer_revenue_rows = []
for cust in ['Food & Groceries', 'LAND Market', "Dominick's"]:
    line = f'Realized revenue - Contracted sales revenue - Contracted sales revenue - {cust}'
    sub = finance_long[finance_long['line_item'] == line][['round', 'value']].copy()
    sub['customer'] = cust
    customer_revenue_rows.append(sub)
customer_revenue = pd.concat(customer_revenue_rows, ignore_index=True)
customer_revenue = customer_revenue.rename(columns={'value': 'contracted_revenue'})

# ---------------- Clean component / product / warehouse tables ----------------
component = component.rename(columns={
    'Component': 'component', 'Round': 'round',
    'Delivery reliability (%)': 'delivery_reliability',
    'Rejection (%)': 'rejection_pct',
    'Component availability (%)': 'component_availability',
    'Obsoletes (%)': 'obsolete_pct',
})
component['round'] = component['round'].astype(int)

product = product.rename(columns={
    'Product': 'product', 'Round': 'round',
    'Service level (pieces)': 'service_level_pieces',
    'Service level (order lines)': 'service_level_order_lines',
    'OSA': 'osa',
    'Obsoletes (%)': 'obsolete_pct',
    'Forecast error (MAPE)': 'mape',
    'Production plan adherence (%)': 'production_plan_adherence',
})
product['round'] = product['round'].astype(int)

customer_product = customer_product.rename(columns={
    'Customer': 'customer', ' Product': 'product', 'Round': 'round',
    'Attained shelf life (%)': 'attained_shelf_life',
    'Service level (pieces)': 'service_level_pieces',
    'Service level (order lines)': 'service_level_order_lines',
    'OSA': 'osa',
})
customer_product['round'] = customer_product['round'].astype(int)

warehouse = warehouse.rename(columns={
    'Warehouse': 'warehouse', 'Round': 'round',
    'Cube utilization (%)': 'cube_utilization',
    'Overflow (%)': 'overflow_pct',
})
warehouse['round'] = warehouse['round'].astype(int)

bottling = bottling.rename(columns={'Round': 'round'})
bottling['round'] = bottling['round'].astype(int)

# ---------------- Write everything out ----------------
financial_kpis.to_csv(f'{OUT}/financial_kpis.csv', index=False)
supplier_purchase.to_csv(f'{OUT}/supplier_purchase.csv', index=False)
customer_bonus.to_csv(f'{OUT}/customer_bonus.csv', index=False)
customer_revenue.to_csv(f'{OUT}/customer_revenue.csv', index=False)
component.to_csv(f'{OUT}/component.csv', index=False)
product.to_csv(f'{OUT}/product.csv', index=False)
customer_product.to_csv(f'{OUT}/customer_product.csv', index=False)
warehouse.to_csv(f'{OUT}/warehouse.csv', index=False)
bottling.to_csv(f'{OUT}/bottling.csv', index=False)

print('Wrote all CSVs to', OUT)
print('financial_kpis:', financial_kpis.shape)
print('component:', component.shape)
print('product:', product.shape)
print('customer_product:', customer_product.shape)
print('warehouse:', warehouse.shape)
